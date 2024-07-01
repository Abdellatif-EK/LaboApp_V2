from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from equipement.models import Equipements

class Command(BaseCommand):
    help = 'Import data from Excel file'

    def add_arguments(self, parser):
        parser.add_argument('file_path', type=str, help='Path to the Excel file')

    def handle(self, *args, **options):
        file_path = options['file_path']
        workbook = load_workbook(filename=file_path)
        sheet = workbook.active
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            equipement = Equipements(
                Emplacement=row[2],
                Appareil=row[3],
                Etat=row[5],
                modele=row[6],
                Code_machine=row[4],
                Password=row[11],
                matrcie_acces=row[16],
                Sauvegarde=row[12],
                Connecte_reseau=row[22],
                Connecte_AD=row[23],
                connecté_imprimante=row[24],
                Maintenance=row[26],
                planning_sauvegarde=row[13],
                Logiciel=row[7],
                version_logiciel=row[17],
                date_installation=row[18],
                Version_windows=row[8],
                Situation=row[9],
                Fournisseur=row[10],
                Etat_materiel_informatique=row[15],
                numero_serie=row[28],
                Documentation=row[19],
                DOC_qualification=row[20],
                QX=row[25],
                description=row[29],
                Num=row[29],
                Autres=row[27]
            )
            equipement.save()
